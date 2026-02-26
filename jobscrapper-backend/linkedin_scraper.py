from __future__ import annotations

import datetime as dt
import os
import re
import time
from html import unescape
from typing import Dict, List, Optional
from urllib.parse import quote_plus, unquote, urlparse

from openpyxl import Workbook


# Runtime-overridable settings (updated by wrapper)
OUTPUT_FILE = "jobs_output.xlsx"
MAX_JOB_AGE_DAYS = 5
HR_KEYWORDS = ["Software Engineer"]
LOCATION_QUERIES = ["India"]

HEADLESS = True
WAIT_SECONDS = 12
SEARCH_PAGES = 1
MAX_RESULTS_PER_QUERY = 20

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)


def _clean_text(value: str) -> str:
    text = unescape(value or "")
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _tokenize(text: str) -> List[str]:
    return [tok for tok in re.split(r"[^a-z0-9]+", (text or "").lower()) if tok]


def _matches_filters(title: str, location: str, role_query: str, location_query: str) -> bool:
    role_tokens = [t for t in _tokenize(role_query) if len(t) >= 2]
    loc_tokens = [t for t in _tokenize(location_query) if len(t) >= 2]

    haystack = title.lower()
    loc_haystack = location.lower()

    if role_tokens and not any(tok in haystack for tok in role_tokens):
        return False
    if loc_tokens and not any(tok in loc_haystack for tok in loc_tokens):
        return False
    return True


def _parse_age_days(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    t = text.lower()
    if any(p in t for p in ["just posted", "today", "new"]):
        return 0
    m = re.search(r"(\d+)\s*minute", t)
    if m:
        return 0
    m = re.search(r"(\d+)\s*hour", t)
    if m:
        return int(m.group(1)) / 24
    m = re.search(r"(\d+)\s*day", t)
    if m:
        return float(int(m.group(1)))
    m = re.search(r"(\d+)\s*week", t)
    if m:
        return float(int(m.group(1)) * 7)
    if "24 hours" in t:
        return 1
    if "last 5 days" in t or "within 5 days" in t:
        return 5
    return None


def _within_age_limit(text: Optional[str], max_days: int = MAX_JOB_AGE_DAYS) -> bool:
    days = _parse_age_days(text)
    if days is None:
        # Keep rows where site does not expose an age label.
        return True
    return days <= max_days


def _decode_yahoo_redirect(url: str) -> str:
    m = re.search(r"/RU=([^/]+)/RK=", url)
    if not m:
        return url
    return unquote(m.group(1))


def _looks_like_listing_url(url: str, portal: str) -> bool:
    u = url.lower()
    if portal == "Indeed":
        return "/jobs?" not in u and "/q-" not in u
    if portal == "Glassdoor":
        return "/job/" in u
    if portal == "Naukri":
        return "job-listings" in u or "-jobs-" in u
    if portal == "Foundit":
        return "foundit" in u and ("job" in u or "jobs" in u)
    return True


def _require_scraper_deps():
    try:
        import requests
        from bs4 import BeautifulSoup
        from selenium import webdriver
        from selenium.common.exceptions import TimeoutException, WebDriverException
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
        from webdriver_manager.chrome import ChromeDriverManager
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Missing Python dependencies for live scraping. Install with: "
            "pip install requests beautifulsoup4 selenium webdriver-manager"
        ) from exc

    return {
        "requests": requests,
        "BeautifulSoup": BeautifulSoup,
        "webdriver": webdriver,
        "TimeoutException": TimeoutException,
        "WebDriverException": WebDriverException,
        "Options": Options,
        "Service": Service,
        "By": By,
        "EC": EC,
        "WebDriverWait": WebDriverWait,
        "ChromeDriverManager": ChromeDriverManager,
    }


def _build_driver(deps):
    options = deps["Options"]()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=en-US")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(f"--user-agent={USER_AGENT}")
    chrome_bin = os.getenv("CHROME_BIN", "").strip()
    chromedriver_path = os.getenv("CHROMEDRIVER_PATH", "").strip()
    if chrome_bin:
        options.binary_location = chrome_bin

    if chromedriver_path:
        service = deps["Service"](chromedriver_path)
    else:
        service = deps["Service"](deps["ChromeDriverManager"]().install())

    return deps["webdriver"].Chrome(service=service, options=options)


def scrape_linkedin_last24h(role_query: str = "", location_query: str = "") -> List[Dict[str, str]]:
    deps = _require_scraper_deps()
    rows: List[Dict[str, str]] = []

    driver = _build_driver(deps)
    try:
        query = role_query or "jobs"
        location = location_query or "India"

        for page in range(SEARCH_PAGES):
            start = page * 25
            url = (
                "https://www.linkedin.com/jobs/search/"
                f"?keywords={quote_plus(query)}&location={quote_plus(location)}"
                f"&f_TPR=r{MAX_JOB_AGE_DAYS * 86400}&start={start}"
            )
            try:
                driver.get(url)
                deps["WebDriverWait"](driver, WAIT_SECONDS).until(
                    deps["EC"].presence_of_element_located((deps["By"].CSS_SELECTOR, "a.base-card__full-link"))
                )
                time.sleep(1.5)
            except (deps["TimeoutException"], deps["WebDriverException"]):
                continue

            soup = deps["BeautifulSoup"](driver.page_source, "html.parser")
            cards = soup.select("li")

            for card in cards:
                a = card.select_one("a.base-card__full-link[href]")
                if not a:
                    continue

                title = _clean_text(a.get_text(" ", strip=True))
                if not title:
                    continue

                company = ""
                for sel in ("h4.base-search-card__subtitle", "a.hidden-nested-link"):
                    n = card.select_one(sel)
                    if n and n.get_text(strip=True):
                        company = _clean_text(n.get_text(" ", strip=True))
                        break

                location_text = ""
                n = card.select_one("span.job-search-card__location")
                if n and n.get_text(strip=True):
                    location_text = _clean_text(n.get_text(" ", strip=True))

                posted_at = ""
                for sel in ("time", "span.job-search-card__listdate", "span.job-search-card__listdate--new"):
                    n = card.select_one(sel)
                    if n and n.get_text(strip=True):
                        posted_at = _clean_text(n.get_text(" ", strip=True))
                        break

                job_url = (a.get("href", "") or "").split("?")[0].strip()

                if not _matches_filters(title, location_text, role_query, location_query):
                    continue
                if not _within_age_limit(posted_at, MAX_JOB_AGE_DAYS):
                    continue

                rows.append(
                    {
                        "title": title,
                        "company": company,
                        "location": location_text,
                        "platform": "LinkedIn",
                        "source": "LinkedIn",
                        "url": job_url,
                        "posted_at": posted_at,
                    }
                )

                if len(rows) >= MAX_RESULTS_PER_QUERY:
                    break

            if len(rows) >= MAX_RESULTS_PER_QUERY:
                break
    finally:
        driver.quit()

    return rows


def yahoo_site_results_last5d(portal_name: str, role_query: str = "", location_query: str = "") -> List[Dict[str, str]]:
    deps = _require_scraper_deps()
    requests = deps["requests"]
    BeautifulSoup = deps["BeautifulSoup"]

    domain_map = {
        "Indeed": "indeed.com",
        "Naukri": "naukri.com",
        "Foundit": "foundit.in",
        "Glassdoor": "glassdoor.com",
    }
    site_query = domain_map.get(portal_name)
    if not site_query:
        return []

    query_parts = [f"site:{site_query}"]
    if role_query.strip():
        query_parts.append(f'"{role_query.strip()}"')
    if location_query.strip():
        query_parts.append(location_query.strip())
    query_parts.append("jobs")
    query_parts.append("last 5 days")
    query = " ".join(query_parts)

    headers = {"User-Agent": USER_AGENT}
    rows: List[Dict[str, str]] = []

    for page in range(SEARCH_PAGES):
        start = page * 10 + 1
        url = f"https://search.yahoo.com/search?p={quote_plus(query)}&b={start}"
        try:
            resp = requests.get(url, headers=headers, timeout=20)
            if resp.status_code >= 400:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
        except requests.RequestException:
            continue

        blocks = soup.select("div#web ol li")
        for block in blocks:
            a = block.select_one("div.compTitle h3 a[href]")
            if not a:
                continue

            title = _clean_text(a.get_text(" ", strip=True))
            raw_url = a.get("href", "").strip()
            real_url = _decode_yahoo_redirect(raw_url)
            netloc = urlparse(real_url).netloc.lower()
            if site_query not in netloc:
                continue
            if not _looks_like_listing_url(real_url, portal_name):
                continue

            snippet_node = block.select_one("div.compText")
            snippet = _clean_text(snippet_node.get_text(" ", strip=True) if snippet_node else "")

            posted_at = "Within 5 days"
            m = re.search(
                r"(\d+\s+(?:hour|hours|day|days)\s+ago|Today|Just posted|last 5 days|within 5 days)",
                snippet,
                flags=re.IGNORECASE,
            )
            if m:
                posted_at = _clean_text(m.group(1))

            location_hint = f"{location_query} {snippet}"
            if not _matches_filters(title, location_hint, role_query, location_query):
                continue
            if not _within_age_limit(posted_at, MAX_JOB_AGE_DAYS):
                continue

            rows.append(
                {
                    "title": title,
                    "company": "",
                    "location": _clean_text(location_query),
                    "platform": portal_name,
                    "source": portal_name,
                    "url": real_url,
                    "posted_at": posted_at,
                }
            )

            if len(rows) >= MAX_RESULTS_PER_QUERY:
                break

        if len(rows) >= MAX_RESULTS_PER_QUERY:
            break

    return rows


def _dedupe(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    unique = {}
    for row in rows:
        key = (row.get("title", "").lower(), row.get("company", "").lower(), row.get("url", ""))
        unique[key] = row
    return list(unique.values())


def _write_xlsx(rows: List[Dict[str, str]], output_file: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["title", "company", "location", "platform", "source", "url", "posted_at"]
    ws.append(headers)

    for row in rows:
        ws.append([row.get(col, "") for col in headers])

    url_col_idx = headers.index("url") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col_idx)
        value = str(cell.value or "").strip()
        if value.startswith("http://") or value.startswith("https://"):
            cell.hyperlink = value
            cell.style = "Hyperlink"

    wb.save(output_file)


def main() -> None:
    rows: List[Dict[str, str]] = []

    roles = HR_KEYWORDS or [""]
    locations = LOCATION_QUERIES or [""]

    for role in roles:
        for location in locations:
            rows.extend(scrape_linkedin_last24h(role, location))
            for portal in ("Naukri", "Indeed", "Foundit", "Glassdoor"):
                rows.extend(yahoo_site_results_last5d(portal, role, location))

    rows = _dedupe(rows)
    _write_xlsx(rows, OUTPUT_FILE)


if __name__ == "__main__":
    main()
