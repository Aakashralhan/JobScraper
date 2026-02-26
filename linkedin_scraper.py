from __future__ import annotations

import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional
from urllib.parse import quote_plus, unquote, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

OUTPUT_FILE = "HR_Jobs_Last24h.xlsx"
HEADLESS = True
WAIT_SECONDS = 12
SEARCH_PAGES = 1
MAX_RESULTS_PER_KEYWORD = 10
LOCATION_QUERIES = [
    "Delhi, India",
    "Noida, Uttar Pradesh, India",
    "Gurugram, Haryana, India",
]
MAX_JOB_AGE_DAYS = 5
TARGET_LOCATION_KEYWORDS = [
    "delhi",
    "new delhi",
    "delhi ncr",
    "noida",
    "gurgaon",
    "gurugram",
    "greater delhi area",
]

HR_KEYWORDS = [
    "HR Manager",
    "HR Executive",
    "HR Generalist",
    "Talent Acquisition",
    "Recruiter",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)


@dataclass
class JobRecord:
    portal: str
    source_keyword: str
    job_title: Optional[str] = None
    company_name: Optional[str] = None
    job_location: Optional[str] = None
    date_posted: Optional[str] = None
    salary_package: Optional[str] = None
    job_url: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    job_description_summary: Optional[str] = None
    employment_type: Optional[str] = None
    fetched_at_utc: Optional[str] = None


def build_driver(headless: bool = True) -> webdriver.Chrome:
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=en-US")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(f"--user-agent={USER_AGENT}")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)


def clean_url(url: str) -> str:
    return url.split("?")[0].strip()


def summarize(text: str, max_chars: int = 550) -> Optional[str]:
    if not text:
        return None
    text = " ".join(text.split())
    return text if len(text) <= max_chars else text[: max_chars - 3] + "..."


def extract_email(text: str) -> Optional[str]:
    if not text:
        return None
    found = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    return found[0] if found else None


def extract_phone(text: str) -> Optional[str]:
    if not text:
        return None
    patterns = [
        r"(?:\+91[-\s]?)?[6-9]\d{9}",
        r"(?:\+\d{1,3}[-\s]?)?\(?\d{2,4}\)?[-\s]?\d{3,5}[-\s]?\d{3,5}",
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return re.sub(r"\s+", " ", m.group(0)).strip()
    return None


def parse_age_days(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    t = text.lower()
    if any(p in t for p in ["just posted", "today", "new"]):
        return 0
    m = re.search(r"(\d+)\s*minute", t)
    if m:
        return 0
    m = re.search(r"(\d+)\s*hour", t)
    if m:
        return int(m.group(1)) / 24
    m = re.search(r"(\d+)\s*day", t)
    if m:
        return float(int(m.group(1)))
    m = re.search(r"(\d+)\s*week", t)
    if m:
        return float(int(m.group(1)) * 7)
    if "24 hours" in t:
        return 1
    if "last 5 days" in t or "within 5 days" in t:
        return 5
    return None


def within_age_limit(text: Optional[str], max_days: int = MAX_JOB_AGE_DAYS) -> bool:
    days = parse_age_days(text)
    if days is None:
        return False
    return days <= max_days


def location_matches(text: Optional[str]) -> bool:
    if not text:
        return False
    t = text.lower()
    return any(k in t for k in TARGET_LOCATION_KEYWORDS)


def looks_last_24h(text: Optional[str]) -> bool:
    days = parse_age_days(text)
    if days is None:
        return False
    return days <= 1


def normalize_tel(phone: str) -> Optional[str]:
    if not phone:
        return None
    digits_plus = re.sub(r"[^\d+]", "", phone)
    if not digits_plus:
        return None
    if digits_plus.startswith("+"):
        return digits_plus
    if len(re.sub(r"\D", "", digits_plus)) == 10:
        return "+91" + re.sub(r"\D", "", digits_plus)
    return "+" + re.sub(r"\D", "", digits_plus)


def apply_excel_hyperlinks(path: str) -> None:
    wb = load_workbook(path)
    ws = wb.active

    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    url_col = header_map.get("job_url")
    email_col = header_map.get("contact_email")
    phone_col = header_map.get("contact_phone")

    for r in range(2, ws.max_row + 1):
        if url_col:
            cell = ws.cell(row=r, column=url_col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

        if email_col:
            cell = ws.cell(row=r, column=email_col)
            if cell.value and isinstance(cell.value, str):
                email = cell.value.strip().split(";")[0].strip()
                if "@" in email:
                    cell.hyperlink = f"mailto:{email}"
                    cell.font = Font(color="0563C1", underline="single")

        if phone_col:
            cell = ws.cell(row=r, column=phone_col)
            if cell.value and isinstance(cell.value, str):
                phone = normalize_tel(cell.value.strip().split(";")[0].strip())
                if phone:
                    cell.hyperlink = f"tel:{phone}"
                    cell.font = Font(color="0563C1", underline="single")

    wb.save(path)


def explain_portal_block(portal: str, text: str) -> None:
    lowered = text.lower()
    if any(k in lowered for k in ["access denied", "cloudflare", "just a moment", "security", "captcha"]):
        print(f"[warn] {portal} appears bot-protected in this environment; partial/zero data possible.")


def looks_like_listing_url(url: str, portal: str) -> bool:
    u = url.lower()
    if portal == "Indeed":
        return "/jobs?" not in u and "/q-" not in u
    if portal == "Glassdoor":
        return "/job/" in u
    if portal == "Naukri":
        return "job-listings" in u or "-jobs-" in u
    return False


def extract_generic_details(url: str) -> Dict[str, Optional[str]]:
    headers = {"User-Agent": USER_AGENT}
    try:
        resp = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        if resp.status_code >= 400:
            return {"salary": None, "email": None, "phone": None, "summary": None, "employment_type": None}
        soup = BeautifulSoup(resp.text, "html.parser")
        body_text = soup.get_text(" ", strip=True)

        salary = None
        salary_patterns = [
            r"\$\s?\d{2,3}[,\d]*(?:\s?-\s?\$\s?\d{2,3}[,\d]*)?",
            r"\d{1,3}[,\d]*\s?(?:USD|INR|per year|per month|LPA)",
        ]
        for pattern in salary_patterns:
            match = re.search(pattern, body_text, flags=re.IGNORECASE)
            if match:
                salary = match.group(0)
                break

        emp = None
        for label in ["Full-time", "Part-time", "Contract", "Internship", "Temporary", "Remote"]:
            if re.search(rf"\b{re.escape(label)}\b", body_text, flags=re.IGNORECASE):
                emp = label
                break

        return {
            "salary": salary,
            "email": extract_email(body_text),
            "phone": extract_phone(body_text),
            "summary": summarize(body_text),
            "employment_type": emp,
        }
    except requests.RequestException:
        return {"salary": None, "email": None, "phone": None, "summary": None, "employment_type": None}


def scrape_linkedin_last24h(driver: webdriver.Chrome, keyword: str, location_query: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    for page in range(SEARCH_PAGES):
        start = page * 25
        url = (
            "https://www.linkedin.com/jobs/search/"
            f"?keywords={quote_plus(keyword)}&location={quote_plus(location_query)}&f_TPR=r432000&start={start}"
        )
        try:
            driver.get(url)
            WebDriverWait(driver, WAIT_SECONDS).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a.base-card__full-link"))
            )
            time.sleep(1.5)
        except TimeoutException:
            continue
        except WebDriverException:
            continue

        soup = BeautifulSoup(driver.page_source, "html.parser")
        cards = soup.select("li")
        count = 0
        for card in cards:
            a = card.select_one("a.base-card__full-link[href]")
            if not a:
                continue
            title = (a.get_text(" ", strip=True) or None)
            company = None
            loc = None
            posted = None
            for sel in ["h4.base-search-card__subtitle", "a.hidden-nested-link"]:
                n = card.select_one(sel)
                if n and n.get_text(strip=True):
                    company = n.get_text(" ", strip=True)
                    break
            for sel in ["span.job-search-card__location", "span.job-search-card__location"]:
                n = card.select_one(sel)
                if n and n.get_text(strip=True):
                    loc = n.get_text(" ", strip=True)
                    break
            for sel in ["time", "span.job-search-card__listdate", "span.job-search-card__listdate--new"]:
                n = card.select_one(sel)
                if n and n.get_text(strip=True):
                    posted = n.get_text(" ", strip=True)
                    break

            job_url = clean_url(a.get("href", ""))
            if not location_matches(loc):
                continue
            if not within_age_limit(posted, max_days=MAX_JOB_AGE_DAYS):
                continue
            details = extract_generic_details(job_url)
            records.append(
                JobRecord(
                    portal="LinkedIn",
                    source_keyword=keyword,
                    job_title=title,
                    company_name=company,
                    job_location=loc,
                    date_posted=posted,
                    salary_package=details["salary"],
                    job_url=job_url,
                    contact_email=details["email"],
                    contact_phone=details["phone"],
                    job_description_summary=details["summary"],
                    employment_type=details["employment_type"],
                    fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                )
            )
            count += 1
            if count >= MAX_RESULTS_PER_KEYWORD:
                break
    return records


def decode_yahoo_redirect(url: str) -> str:
    # Example redirect:
    # https://r.search.yahoo.com/.../RU=https%3a%2f%2fwww.indeed.com%2f.../RK=...
    m = re.search(r"/RU=([^/]+)/RK=", url)
    if not m:
        return url
    return unquote(m.group(1))


def yahoo_site_results_last5d(portal_name: str, site_query: str, keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    query = (
        f'site:{site_query} "{keyword}" jobs '
        '"Delhi NCR" OR "Noida" OR "Gurgaon" OR "Gurugram" "last 5 days"'
    )
    headers = {"User-Agent": USER_AGENT}

    for page in range(SEARCH_PAGES):
        start = page * 10 + 1
        url = f"https://search.yahoo.com/search?p={quote_plus(query)}&b={start}"
        try:
            resp = requests.get(url, headers=headers, timeout=20)
            if resp.status_code >= 400:
                print(
                    f"[warn] Yahoo site search failed for {portal_name} page={page + 1}: "
                    f"status={resp.status_code} {resp.reason}"
                )
                continue
            explain_portal_block(portal_name, resp.text[:6000])
            soup = BeautifulSoup(resp.text, "html.parser")
        except requests.RequestException as exc:
            print(f"[warn] Yahoo site search request failed for {portal_name}: {exc}")
            continue

        blocks = soup.select("div#web ol li")
        for block in blocks:
            a = block.select_one("div.compTitle h3 a[href]")
            if not a:
                continue
            title = a.get_text(" ", strip=True) or None
            raw_url = a.get("href", "").strip()
            real_url = decode_yahoo_redirect(raw_url)
            netloc = urlparse(real_url).netloc
            if site_query not in netloc:
                continue
            if not looks_like_listing_url(real_url, portal_name):
                continue

            snippet_node = block.select_one("div.compText")
            snippet = snippet_node.get_text(" ", strip=True) if snippet_node else ""
            if not location_matches(f"{title or ''} {snippet}"):
                continue

            date_posted = None
            m = re.search(
                r"(\d+\s+(?:hour|hours|day|days)\s+ago|Today|Just posted|last 5 days|within 5 days)",
                snippet,
                flags=re.IGNORECASE,
            )
            if m:
                date_posted = m.group(1)
            else:
                date_posted = "Within 5 days (search-filtered)"
            if not within_age_limit(date_posted, max_days=MAX_JOB_AGE_DAYS):
                continue

            details = extract_generic_details(real_url)
            records.append(
                JobRecord(
                    portal=portal_name,
                    source_keyword=keyword,
                    job_title=title,
                    company_name=None,
                    job_location=None,
                    date_posted=date_posted,
                    salary_package=details["salary"],
                    job_url=real_url,
                    contact_email=details["email"],
                    contact_phone=details["phone"],
                    job_description_summary=details["summary"] or summarize(snippet),
                    employment_type=details["employment_type"],
                    fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                )
            )
            if len(records) >= MAX_RESULTS_PER_KEYWORD:
                break
        if len(records) >= MAX_RESULTS_PER_KEYWORD:
            break

    return records


def to_dataframe(records: List[JobRecord]) -> pd.DataFrame:
    rows = [asdict(r) for r in records]
    cols = [
        "portal",
        "source_keyword",
        "job_title",
        "company_name",
        "job_location",
        "date_posted",
        "salary_package",
        "job_url",
        "contact_email",
        "contact_phone",
        "job_description_summary",
        "employment_type",
        "fetched_at_utc",
    ]
    if not rows:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(rows)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols]


def main() -> None:
    print("[info] Starting HR jobs scrape for Delhi NCR/Noida/Gurgaon (last 24h to 5 days)")
    driver = build_driver(headless=HEADLESS)
    all_records: List[JobRecord] = []

    try:
        for kw in HR_KEYWORDS:
            for locq in LOCATION_QUERIES:
                print(f"[info] LinkedIn: {kw} | {locq}")
                all_records.extend(scrape_linkedin_last24h(driver, kw, locq))

            print(f"[info] Indeed last-5-days discovery: {kw}")
            all_records.extend(yahoo_site_results_last5d("Indeed", "indeed.com", kw))

            print(f"[info] Naukri last-5-days discovery: {kw}")
            all_records.extend(yahoo_site_results_last5d("Naukri", "naukri.com", kw))

            print(f"[info] Glassdoor last-5-days discovery: {kw}")
            all_records.extend(yahoo_site_results_last5d("Glassdoor", "glassdoor.com", kw))

            time.sleep(1)
    finally:
        driver.quit()

    df = to_dataframe(all_records)
    if not df.empty:
        df.drop_duplicates(subset=["portal", "job_url"], inplace=True)
    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")
    apply_excel_hyperlinks(OUTPUT_FILE)
    print(f"[done] Saved {len(df)} records to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
